"""
generate_report.py
==================
Digipay E2E Test Suite — XLSX Report Generator

Reads pytest's JSON output file (generated with --json-report plugin)
and produces a professional multi-sheet XLSX report.

Usage (called by pytest conftest plugin OR directly):
    python generate_report.py --input .test-results/report.json
                              --output reports/E2E_Test_Report_Digipay.xlsx

The report mirrors the PancreaScan sample structure:
  Sheet 1 – Summary         (pass/fail counts, duration, environment)
  Sheet 2 – Test Cases      (full test-by-test breakdown)
  Sheet 3 – Failures Detail (traceback for failed tests)
  Sheet 4 – Category Pivot  (grouped by module / category)
"""

import os
import sys
import json
import argparse
from datetime import datetime, timezone

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataLabelList
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
# Colour palette (matches Digipay brand — dark blue / cyan / slate)
# ─────────────────────────────────────────────────────────────────────────────
class Colors:
    HEADER_BG      = "1E3A5F"   # Deep navy blue — header row
    HEADER_FG      = "FFFFFF"   # White text
    PASS_BG        = "D1FAE5"   # Soft green
    PASS_FG        = "065F46"   # Dark green
    FAIL_BG        = "FEE2E2"   # Soft red
    FAIL_FG        = "991B1B"   # Dark red
    SKIP_BG        = "FEF3C7"   # Soft amber
    SKIP_FG        = "92400E"   # Dark amber
    ERROR_BG       = "FDE8D8"   # Soft orange
    ERROR_FG       = "9A3412"   # Dark orange
    TITLE_BG       = "0F2942"   # Very dark navy (title rows)
    TITLE_FG       = "06B6D4"   # Brand cyan
    SECTION_BG     = "EFF6FF"   # Very light blue
    ALT_ROW        = "F8FAFC"   # Light slate for alternating rows
    BORDER_COLOR   = "CBD5E1"   # Slate border
    CATEGORY_BG    = "DBEAFE"   # Light blue for category rows
    SUMMARY_METRIC = "1D4ED8"   # Brand blue for metric values


# ─────────────────────────────────────────────────────────────────────────────
# Style helpers
# ─────────────────────────────────────────────────────────────────────────────
def make_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def make_border(color: str = Colors.BORDER_COLOR) -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def apply_header_style(cell, bg: str = Colors.HEADER_BG, fg: str = Colors.HEADER_FG,
                        bold: bool = True, size: int = 10):
    cell.fill = make_fill(bg)
    cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()


def apply_data_style(cell, bg: str = "FFFFFF", fg: str = "1E293B",
                      bold: bool = False, align: str = "left", size: int = 9):
    cell.fill = make_fill(bg)
    cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = make_border()


def status_colors(status: str):
    s = (status or "").upper()
    if s == "PASSED":
        return Colors.PASS_BG, Colors.PASS_FG
    elif s == "FAILED":
        return Colors.FAIL_BG, Colors.FAIL_FG
    elif s in ("SKIPPED", "XFAILED"):
        return Colors.SKIP_BG, Colors.SKIP_FG
    else:
        return Colors.ERROR_BG, Colors.ERROR_FG


def set_col_width(ws, col_idx: int, width: float):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


# ─────────────────────────────────────────────────────────────────────────────
# Parse JSON report
# ─────────────────────────────────────────────────────────────────────────────
def load_json_report(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def extract_tests(report: dict) -> list[dict]:
    """Flatten pytest-json-report structure into list of test dicts."""
    tests = []
    for t in report.get("tests", []):
        node_id    = t.get("nodeid", "")
        module_raw = node_id.split("::")[0].replace("tests/", "").replace(".py", "")
        test_name  = "::".join(node_id.split("::")[1:]) if "::" in node_id else node_id
        outcome    = t.get("outcome", "unknown").upper()
        duration   = round(t.get("duration", 0.0), 3)

        # Extract docstring if present
        metadata   = t.get("metadata", {}) or {}
        call_phase = t.get("call", {}) or {}
        longrepr   = call_phase.get("longrepr", "") if isinstance(call_phase, dict) else ""
        if not isinstance(longrepr, str):
            longrepr = str(longrepr)

        # Determine category from module name
        category_map = {
            "test_landing_page": "Functionality",
            "test_login_portal": "Functionality",
            "test_dashboard":    "Functionality",
            "test_vulnerability":"Vulnerability",
            "test_unit":         "Unit",
        }
        category = category_map.get(module_raw.split("/")[-1], "Other")

        # Extract TC number from test function name
        tc_num = ""
        if "_tc" in test_name.lower():
            parts = test_name.lower().split("_tc")
            if len(parts) > 1:
                num_part = parts[1].split("_")[0]
                if num_part.isdigit():
                    tc_num = f"TC-{num_part.zfill(3)}"

        tests.append({
            "tc_id":       tc_num or node_id[:12],
            "module":      module_raw.replace("tests/", "").replace("test_", "").replace("_", " ").title(),
            "category":    category,
            "class":       node_id.split("::")[1] if len(node_id.split("::")) > 2 else "",
            "test_name":   test_name,
            "description": metadata.get("description", test_name.replace("_", " ").title()),
            "status":      outcome,
            "duration_s":  duration,
            "failure_msg": longrepr[:2000] if longrepr else "",
        })
    return tests


def build_fallback_tests() -> list[dict]:
    """If no JSON report exists, return the static test-case registry."""
    registry = [
        # Landing Page
        ("TC-001","Landing Page","Functionality","TestLandingPageLoad","test_tc001_page_loads_successfully","Landing page loads with HTTP 200 and body is visible"),
        ("TC-002","Landing Page","Functionality","TestLandingPageLoad","test_tc002_page_title_contains_digipay","Document title includes Digipay branding"),
        ("TC-003","Landing Page","Functionality","TestLandingPageLoad","test_tc003_no_javascript_errors_on_load","Browser console has no critical JS errors on load"),
        ("TC-004","Landing Page","Functionality","TestLandingPageLoad","test_tc004_page_source_not_empty","Page source is not empty"),
        ("TC-005","Landing Page","Functionality","TestLandingPageLoad","test_tc005_favicon_present","Favicon link tag is present in head"),
        ("TC-006","Landing Page","Functionality","TestLandingPageHeader","test_tc006_header_is_visible","Sticky header element is rendered and visible"),
        ("TC-007","Landing Page","Functionality","TestLandingPageHeader","test_tc007_brand_logo_text_visible","DIGIPAY brand name is shown in the header"),
        ("TC-008","Landing Page","Functionality","TestLandingPageHeader","test_tc008_signin_button_visible","Sign In nav button is present and visible"),
        ("TC-009","Landing Page","Functionality","TestLandingPageHeader","test_tc009_signin_button_text","Sign In button displays correct label text"),
        ("TC-010","Landing Page","Functionality","TestLandingPageHeader","test_tc010_go_to_console_button_visible","Go to Console admin button is visible in header"),
        ("TC-011","Landing Page","Functionality","TestLandingPageHeader","test_tc011_header_is_sticky","Header has sticky/fixed CSS positioning"),
        ("TC-012","Landing Page","Functionality","TestLandingPageHeader","test_tc012_header_backdrop_blur_applied","Header uses backdrop blur for glassmorphism effect"),
        ("TC-013","Landing Page","Functionality","TestLandingPageHeroSection","test_tc013_hero_h1_is_present","h1 heading exists in hero section"),
        ("TC-014","Landing Page","Functionality","TestLandingPageHeroSection","test_tc014_hero_heading_content","Hero heading contains UPI Payments text"),
        ("TC-015","Landing Page","Functionality","TestLandingPageHeroSection","test_tc015_hero_subtext_visible","Hero description paragraph is visible"),
        ("TC-016","Landing Page","Functionality","TestLandingPageHeroSection","test_tc016_open_portal_button_present","Open Web Portal CTA button is present"),
        ("TC-017","Landing Page","Functionality","TestLandingPageHeroSection","test_tc017_open_portal_button_text","Main CTA button text says Open Web Portal"),
        ("TC-018","Landing Page","Functionality","TestLandingPageHeroSection","test_tc018_launch_ios_app_link_present","Launch iOS App deep-link button is visible"),
        ("TC-019","Landing Page","Functionality","TestLandingPageHeroSection","test_tc019_hero_card_display_visible","Hero interactive UI card panel is rendered"),
        ("TC-020","Landing Page","Functionality","TestLandingPageHeroSection","test_tc020_badge_text_visible","Next-Gen QR-less Payment System badge is visible"),
        ("TC-021","Landing Page","Functionality","TestLandingPageHeroSection","test_tc021_open_portal_click_navigates","Clicking Open Web Portal navigates to login page"),
        ("TC-022","Landing Page","Functionality","TestLandingPageHeroSection","test_tc022_signin_nav_click_navigates","Clicking Sign In navbar button opens login portal"),
        ("TC-023","Landing Page","Functionality","TestLandingPageFeatures","test_tc023_features_section_present","Core features section with h2 heading is visible"),
        ("TC-024","Landing Page","Functionality","TestLandingPageFeatures","test_tc024_digipin_feature_card_visible","DIGIPIN Address Translation feature card is present"),
        ("TC-025","Landing Page","Functionality","TestLandingPageFeatures","test_tc025_heading_speed_scoring_visible","Heading and Speed Scoring feature card is present"),
        ("TC-026","Landing Page","Functionality","TestLandingPageFeatures","test_tc026_autonomous_categorization_visible","Autonomous Categorization feature card is present"),
        ("TC-027","Landing Page","Functionality","TestLandingPageFeatures","test_tc027_download_section_present","App download install banner section is rendered"),
        ("TC-028","Landing Page","Functionality","TestLandingPageFeatures","test_tc028_qr_code_graphic_present","SVG QR code graphic is rendered"),
        ("TC-029","Landing Page","Functionality","TestLandingPageFeatures","test_tc029_footer_visible","Footer element is present at bottom"),
        ("TC-030","Landing Page","Functionality","TestLandingPageFeatures","test_tc030_footer_copyright_text","Footer contains copyright year and DIGIPAY text"),
        # Login Portal
        ("TC-031","Login Portal","Functionality","TestLoginNavigation","test_tc031_login_portal_renders","Login portal page renders after clicking Sign In"),
        ("TC-032","Login Portal","Functionality","TestLoginNavigation","test_tc032_back_button_visible","Back to Home button is visible on login portal"),
        ("TC-033","Login Portal","Functionality","TestLoginNavigation","test_tc033_back_button_navigates_to_landing","Clicking back button returns user to landing page"),
        ("TC-034","Login Portal","Functionality","TestLoginNavigation","test_tc034_login_portal_heading_visible","Secure Login Portal heading is visible"),
        ("TC-035","Login Portal","Functionality","TestLoginNavigation","test_tc035_country_code_displayed","+91 India country code prefix is visible"),
        ("TC-036","Login Portal","Functionality","TestLoginNavigation","test_tc036_admin_hint_displayed","Admin test phone hint 9999999999 is shown"),
        ("TC-037","Login Portal","Functionality","TestPhoneInputValidation","test_tc037_phone_input_is_present","Phone number input field is present and enabled"),
        ("TC-038","Login Portal","Functionality","TestPhoneInputValidation","test_tc038_phone_input_type_tel","Phone input has type tel for mobile keyboard"),
        ("TC-039","Login Portal","Functionality","TestPhoneInputValidation","test_tc039_phone_max_length_10","Phone input enforces maxlength of 10 digits"),
        ("TC-040","Login Portal","Functionality","TestPhoneInputValidation","test_tc040_phone_accepts_digits","Entering digits into phone field is accepted"),
        ("TC-041","Login Portal","Functionality","TestPhoneInputValidation","test_tc041_phone_rejects_letters","Alphabetic characters are filtered out from phone input"),
        ("TC-042","Login Portal","Functionality","TestPhoneInputValidation","test_tc042_empty_phone_shows_toast","Submitting with empty phone shows validation toast"),
        ("TC-043","Login Portal","Functionality","TestPhoneInputValidation","test_tc043_short_phone_shows_validation","Phone number shorter than 10 digits is rejected"),
        ("TC-044","Login Portal","Functionality","TestPhoneInputValidation","test_tc044_request_code_button_present","Request Code submit button is visible and clickable"),
        ("TC-045","Login Portal","Functionality","TestPhoneInputValidation","test_tc045_phone_placeholder_visible","Placeholder text is shown in empty phone input"),
        ("TC-046","Login Portal","Functionality","TestOTPFlow","test_tc046_valid_phone_triggers_otp_step","Valid 10-digit phone triggers OTP input step"),
        ("TC-047","Login Portal","Functionality","TestOTPFlow","test_tc047_otp_input_renders_after_phone_submit","OTP 6-digit input field appears after valid submit"),
        ("TC-048","Login Portal","Functionality","TestOTPFlow","test_tc048_otp_input_type_text","OTP input has type text for 6-digit entry"),
        ("TC-049","Login Portal","Functionality","TestOTPFlow","test_tc049_otp_max_length_6","OTP input enforces maxlength of 6"),
        ("TC-050","Login Portal","Functionality","TestOTPFlow","test_tc050_dev_otp_hint_block_shown","Developer OTP hint block appears after requesting code"),
        ("TC-051","Login Portal","Functionality","TestOTPFlow","test_tc051_change_phone_number_link_visible","Change Phone Number link is visible on OTP step"),
        ("TC-052","Login Portal","Functionality","TestOTPFlow","test_tc052_change_phone_returns_to_step1","Clicking Change Phone Number returns to phone step"),
        ("TC-053","Login Portal","Functionality","TestOTPFlow","test_tc053_invalid_otp_shows_error","Entering wrong 6-digit OTP shows error message"),
        ("TC-054","Login Portal","Functionality","TestOTPFlow","test_tc054_otp_filters_non_digits","OTP input strips non-numeric characters"),
        ("TC-055","Login Portal","Functionality","TestOTPFlow","test_tc055_loading_spinner_shown_during_otp_request","Loading state appears while OTP is being requested"),
        ("TC-056","Login Portal","Functionality","TestOTPFlow","test_tc056_short_otp_blocked","OTP shorter than 6 digits is rejected on submit"),
        ("TC-057","Login Portal","Functionality","TestOTPFlow","test_tc057_verify_continue_button_label","OTP step verify button shows Verify and Continue text"),
        ("TC-058","Login Portal","Functionality","TestOTPFlow","test_tc058_toast_component_dismissible","Error toast notification can be dismissed"),
        ("TC-059","Login Portal","Functionality","TestOTPFlow","test_tc059_login_page_no_sensitive_data_in_url","Phone number or OTP not exposed in browser URL"),
        ("TC-060","Login Portal","Functionality","TestOTPFlow","test_tc060_back_from_otp_clears_state","Navigating back from OTP step clears OTP field"),
        # Dashboard
        ("TC-061","Dashboard","Functionality","TestDashboardNavBar","test_tc061_dashboard_navbar_visible_after_login","Dashboard nav renders after successful authentication"),
        ("TC-062","Dashboard","Functionality","TestDashboardNavBar","test_tc062_dashboard_brand_logo_present","DIGIPAY CONSOLE brand name appears in dashboard nav"),
        ("TC-063","Dashboard","Functionality","TestDashboardNavBar","test_tc063_logout_button_has_correct_testid","Logout button data-testid exists in DOM"),
        ("TC-064","Dashboard","Functionality","TestDashboardNavBar","test_tc064_logout_clears_session","Clicking logout removes token from localStorage"),
        ("TC-065","Dashboard","Functionality","TestDashboardNavBar","test_tc065_landing_page_after_logout","After logout user is redirected to landing page"),
        ("TC-066","Dashboard","Functionality","TestDashboardNavBar","test_tc066_expired_token_redirects_to_landing","Expired JWT token redirects to landing"),
        ("TC-067","Dashboard","Functionality","TestDashboardNavBar","test_tc067_no_token_shows_landing","Without stored token landing page is shown"),
        ("TC-068","Dashboard","Functionality","TestDashboardNavBar","test_tc068_dashboard_loading_spinner","Loading spinner shown while dashboard fetches data"),
        ("TC-069","Dashboard","Functionality","TestAdminDashboardTabs","test_tc069_admin_overview_tab_testid_present","Admin overview tab testid present"),
        ("TC-070","Dashboard","Functionality","TestAdminDashboardTabs","test_tc070_admin_transactions_tab_testid","Transactions tab testid correct"),
        ("TC-071","Dashboard","Functionality","TestAdminDashboardTabs","test_tc071_admin_merchants_tab_testid","Merchants tab testid correct"),
        ("TC-072","Dashboard","Functionality","TestAdminDashboardTabs","test_tc072_admin_analytics_tab_testid","Analytics tab testid correct"),
        ("TC-073","Dashboard","Functionality","TestAdminDashboardTabs","test_tc073_search_input_testid_present","Transaction search input testid present"),
        ("TC-074","Dashboard","Functionality","TestAdminDashboardTabs","test_tc074_category_filter_testid_present","Category filter select testid present"),
        ("TC-075","Dashboard","Functionality","TestAdminDashboardTabs","test_tc075_export_csv_btn_testid_present","Export CSV button testid present"),
        ("TC-076","Dashboard","Functionality","TestDashboardDataDisplay","test_tc076_customer_dashboard_balance_kpi","Customer dashboard shows Estimated Balance KPI card"),
        ("TC-077","Dashboard","Functionality","TestDashboardDataDisplay","test_tc077_merchant_dashboard_revenue_kpi","Merchant dashboard shows Today Revenue KPI"),
        ("TC-078","Dashboard","Functionality","TestDashboardDataDisplay","test_tc078_admin_kpi_today_revenue","Admin overview shows Today Revenue KPI card"),
        ("TC-079","Dashboard","Functionality","TestDashboardDataDisplay","test_tc079_admin_kpi_total_transactions","Admin overview shows Total Transactions KPI"),
        ("TC-080","Dashboard","Functionality","TestDashboardDataDisplay","test_tc080_admin_kpi_registered_users","Admin overview shows Registered Users KPI"),
        ("TC-081","Dashboard","Functionality","TestDashboardDataDisplay","test_tc081_admin_kpi_total_merchants","Admin overview shows Total Merchants KPI"),
        ("TC-082","Dashboard","Functionality","TestDashboardDataDisplay","test_tc082_transaction_table_headers","Transaction ledger table shows correct column headers"),
        ("TC-083","Dashboard","Functionality","TestDashboardDataDisplay","test_tc083_category_filter_options","Category filter includes Food Shopping Medical Bills"),
        ("TC-084","Dashboard","Functionality","TestDashboardDataDisplay","test_tc084_csv_export_creates_download","CSV export button triggers file download"),
        ("TC-085","Dashboard","Functionality","TestDashboardDataDisplay","test_tc085_pagination_controls_visible","Prev Next pagination controls appear when data > 10 rows"),
        # Vulnerability
        ("TC-086","Vulnerability","Vulnerability","TestXSSVulnerabilities","test_tc086_xss_payload_in_phone_field_not_executed","XSS script payload in phone input is not executed"),
        ("TC-087","Vulnerability","Vulnerability","TestXSSVulnerabilities","test_tc087_xss_payload_stripped_in_otp_field","XSS payload in OTP field is stripped"),
        ("TC-088","Vulnerability","Vulnerability","TestXSSVulnerabilities","test_tc088_no_alert_dialogs_on_xss_attempt","Browser alert dialog does NOT appear on XSS attempts"),
        ("TC-089","Vulnerability","Vulnerability","TestXSSVulnerabilities","test_tc089_page_source_does_not_reflect_xss_input","XSS input not reflected unescaped in page source"),
        ("TC-090","Vulnerability","Vulnerability","TestXSSVulnerabilities","test_tc090_react_jsx_escapes_user_input","React JSX auto-escapes any text rendered from state"),
        ("TC-091","Vulnerability","Vulnerability","TestHTTPSTransportSecurity","test_tc091_app_served_over_https","Application URL uses HTTPS protocol"),
        ("TC-092","Vulnerability","Vulnerability","TestHTTPSTransportSecurity","test_tc092_backend_url_uses_https","Backend API base URL uses HTTPS"),
        ("TC-093","Vulnerability","Vulnerability","TestHTTPSTransportSecurity","test_tc093_no_mixed_content_in_page","No mixed HTTP HTTPS content errors in browser console"),
        ("TC-094","Vulnerability","Vulnerability","TestHTTPSTransportSecurity","test_tc094_hsts_header_or_github_pages_enforced","GitHub Pages enforces HTTPS by default"),
        ("TC-095","Vulnerability","Vulnerability","TestAuthorizationSecurity","test_tc095_direct_url_without_auth_lands_on_landing","Without a token shows landing page not dashboard"),
        ("TC-096","Vulnerability","Vulnerability","TestAuthorizationSecurity","test_tc096_malformed_jwt_rejected","Malformed JWT token leads to landing page"),
        ("TC-097","Vulnerability","Vulnerability","TestAuthorizationSecurity","test_tc097_token_stored_in_localstorage_not_cookie","JWT token stored in localStorage not exposed as cookie"),
        ("TC-098","Vulnerability","Vulnerability","TestAuthorizationSecurity","test_tc098_phone_number_not_in_page_source_after_submit","Admin phone not exposed unmasked in DOM after OTP step"),
        ("TC-099","Vulnerability","Vulnerability","TestInjectionAndClickjacking","test_tc099_sql_injection_in_phone_field_sanitised","SQL injection string in phone field is sanitised"),
        ("TC-100","Vulnerability","Vulnerability","TestInjectionAndClickjacking","test_tc100_script_tag_in_search_field_blocked","Script injection in search inputs does not execute"),
        ("TC-101","Vulnerability","Vulnerability","TestInjectionAndClickjacking","test_tc101_open_redirect_not_possible_via_url","App does not redirect to external URLs via URL manipulation"),
        ("TC-102","Vulnerability","Vulnerability","TestInjectionAndClickjacking","test_tc102_clickjacking_x_frame_options_header","Page cannot be embedded in an iframe"),
        # Unit
        ("TC-103","Unit","Unit","TestComponentAttributes","test_tc103_login_nav_button_testid","data-testid login-nav-button present on Sign In nav button"),
        ("TC-104","Unit","Unit","TestComponentAttributes","test_tc104_login_nav_admin_testid","data-testid login-nav-admin present on Go to Console button"),
        ("TC-105","Unit","Unit","TestComponentAttributes","test_tc105_hero_login_button_testid","data-testid login-button present on hero CTA button"),
        ("TC-106","Unit","Unit","TestComponentAttributes","test_tc106_open_app_link_testid","data-testid open-app-link present on Launch iOS App anchor"),
        ("TC-107","Unit","Unit","TestComponentAttributes","test_tc107_open_app_link_href","iOS app deep link has href=digipay://"),
        ("TC-108","Unit","Unit","TestComponentAttributes","test_tc108_phone_input_testid","data-testid phone-input present on phone number input"),
        ("TC-109","Unit","Unit","TestComponentAttributes","test_tc109_back_button_testid","data-testid back-button present on Back to Home button"),
        ("TC-110","Unit","Unit","TestComponentAttributes","test_tc110_login_submit_button_testid","data-testid login-button present on Request Code submit"),
        ("TC-111","Unit","Unit","TestDOMStructure","test_tc111_single_h1_on_landing","Exactly one h1 element on the landing page"),
        ("TC-112","Unit","Unit","TestDOMStructure","test_tc112_main_element_present","Semantic main element is present on landing page"),
        ("TC-113","Unit","Unit","TestDOMStructure","test_tc113_nav_element_present","Semantic nav or header element is used for navigation"),
        ("TC-114","Unit","Unit","TestDOMStructure","test_tc114_section_elements_for_content","Semantic section elements used for content blocks"),
        ("TC-115","Unit","Unit","TestDOMStructure","test_tc115_no_inline_script_in_body","No dangerous inline script tags are injected in body"),
        ("TC-116","Unit","Unit","TestDOMStructure","test_tc116_all_images_have_alt_text","All img elements have non-empty alt attributes"),
        ("TC-117","Unit","Unit","TestResponsivenessAndViewport","test_tc117_mobile_viewport_renders","App renders correctly at mobile viewport 375x812"),
        ("TC-118","Unit","Unit","TestResponsivenessAndViewport","test_tc118_tablet_viewport_renders","App renders correctly at tablet viewport 768x1024"),
        ("TC-119","Unit","Unit","TestResponsivenessAndViewport","test_tc119_desktop_viewport_renders","App renders correctly at full desktop viewport 1920x1080"),
        ("TC-120","Unit","Unit","TestResponsivenessAndViewport","test_tc120_no_horizontal_scrollbar_on_desktop","No horizontal scrollbar appears at desktop viewport"),
    ]
    return [
        {
            "tc_id": r[0], "module": r[1], "category": r[2],
            "class": r[3], "test_name": r[4], "description": r[5],
            "status": "PENDING", "duration_s": 0.0, "failure_msg": ""
        }
        for r in registry
    ]


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1: Summary
# ─────────────────────────────────────────────────────────────────────────────
def write_summary_sheet(wb: openpyxl.Workbook, tests: list[dict], run_meta: dict):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 10

    total    = len(tests)
    passed   = sum(1 for t in tests if t["status"] == "PASSED")
    failed   = sum(1 for t in tests if t["status"] == "FAILED")
    skipped  = sum(1 for t in tests if t["status"] in ("SKIPPED", "XFAILED"))
    errors   = sum(1 for t in tests if t["status"] in ("ERROR", "PENDING", "UNKNOWN"))
    duration = sum(t["duration_s"] for t in tests)
    pass_pct = round((passed / total * 100) if total else 0, 1)

    # ── Title Banner ──────────────────────────────────────────────────────────
    ws.merge_cells("B2:I2")
    title_cell = ws["B2"]
    title_cell.value = "🧪  DIGIPAY  —  E2E TEST EXECUTION REPORT"
    title_cell.fill  = make_fill(Colors.TITLE_BG)
    title_cell.font  = Font(bold=True, color=Colors.TITLE_FG, size=16, name="Calibri")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 38

    ws.merge_cells("B3:I3")
    sub_cell = ws["B3"]
    ts = run_meta.get("timestamp", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"))
    sub_cell.value = f"Generated: {ts}  |  Environment: {run_meta.get('env', 'GitHub Pages + Railway')}  |  Framework: Selenium + pytest"
    sub_cell.fill  = make_fill("0F2942")
    sub_cell.font  = Font(color="94A3B8", size=9, italic=True, name="Calibri")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20

    # ── KPI Cards ─────────────────────────────────────────────────────────────
    kpis = [
        ("Total Test Cases", str(total),    "1E3A5F", "FFFFFF"),
        ("✅  Passed",        str(passed),   "065F46", "D1FAE5"),
        ("❌  Failed",        str(failed),   "991B1B", "FEE2E2"),
        ("⚠️  Skipped",       str(skipped),  "92400E", "FEF3C7"),
        ("Pass Rate",         f"{pass_pct}%","1D4ED8", "EFF6FF"),
        ("Duration",          f"{duration:.1f}s", "374151", "F1F5F9"),
    ]

    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 40
    ws.row_dimensions[7].height = 22

    for col_offset, (label, value, fg, bg) in enumerate(kpis):
        col = col_offset + 2
        # Label
        lc = ws.cell(row=5, column=col)
        lc.value = label
        lc.fill  = make_fill(bg)
        lc.font  = Font(bold=True, color=fg, size=9, name="Calibri")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = make_border()
        # Value
        vc = ws.cell(row=6, column=col)
        vc.value = value
        vc.fill  = make_fill(bg)
        vc.font  = Font(bold=True, color=fg, size=22, name="Calibri")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = make_border()

    # ── Run Metadata ──────────────────────────────────────────────────────────
    meta_rows = [
        ("Run ID",        run_meta.get("run_id",       "N/A")),
        ("Repository",    run_meta.get("repo",         "harishbalaji826-ops/Digipay-Web")),
        ("Branch",        run_meta.get("branch",       "main")),
        ("Commit SHA",    run_meta.get("sha",          "N/A")),
        ("App URL",       run_meta.get("app_url",      "https://harishbalaji826-ops.github.io/Digipay-Web/")),
        ("Backend URL",   run_meta.get("backend_url",  "Railway")),
        ("Browser",       run_meta.get("browser",      "Chrome Headless")),
        ("Test Runner",   "pytest + Selenium 4"),
        ("Python",        run_meta.get("python",       sys.version.split()[0])),
        ("Report Date",   ts),
    ]

    ws.row_dimensions[9].height = 20
    h_cell = ws["B9"]
    h_cell.value = "⚙️  Execution Environment"
    h_cell.fill  = make_fill(Colors.HEADER_BG)
    h_cell.font  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    h_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("B9:I9")

    for i, (k, v) in enumerate(meta_rows, start=10):
        row = i
        bg = Colors.ALT_ROW if i % 2 == 0 else "FFFFFF"
        kc = ws.cell(row=row, column=2, value=k)
        apply_data_style(kc, bg=Colors.SECTION_BG, bold=True, align="left")
        ws.merge_cells(f"B{row}:C{row}")
        vc = ws.cell(row=row, column=4, value=v)
        apply_data_style(vc, bg=bg, align="left")
        ws.merge_cells(f"D{row}:I{row}")
        ws.row_dimensions[row].height = 18

    # ── Category Breakdown ────────────────────────────────────────────────────
    cats = {}
    for t in tests:
        c = t["category"]
        if c not in cats:
            cats[c] = {"total": 0, "passed": 0, "failed": 0}
        cats[c]["total"]  += 1
        cats[c]["passed"] += 1 if t["status"] == "PASSED" else 0
        cats[c]["failed"] += 1 if t["status"] == "FAILED" else 0

    start_row = 22
    ws.row_dimensions[start_row].height = 20
    cat_title = ws.cell(row=start_row, column=2)
    cat_title.value = "📊  Results by Category"
    cat_title.fill  = make_fill(Colors.HEADER_BG)
    cat_title.font  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    cat_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"B{start_row}:I{start_row}")

    headers = ["Category", "Total", "Passed", "Failed", "Pass Rate"]
    for ci, h in enumerate(headers, start=2):
        cell = ws.cell(row=start_row + 1, column=ci)
        apply_header_style(cell)
        cell.value = h
        ws.row_dimensions[start_row + 1].height = 18

    for ri, (cat_name, counts) in enumerate(cats.items(), start=start_row + 2):
        pct = round(counts["passed"] / counts["total"] * 100, 1) if counts["total"] else 0
        row_data = [cat_name, counts["total"], counts["passed"], counts["failed"], f"{pct}%"]
        bg = Colors.ALT_ROW if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row_data, start=2):
            cell = ws.cell(row=ri, column=ci, value=val)
            apply_data_style(cell, bg=bg, align="center" if ci > 2 else "left",
                              bold=(ci == 2))
        ws.row_dimensions[ri].height = 17

    # Column widths
    for col, width in zip(range(2, 10), [20, 14, 14, 14, 14, 16, 18, 20]):
        set_col_width(ws, col, width)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2: Test Cases
# ─────────────────────────────────────────────────────────────────────────────
def write_testcases_sheet(wb: openpyxl.Workbook, tests: list[dict]):
    ws = wb.create_sheet("Test Cases")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = [
        "TC ID", "Module", "Category", "Class / Suite",
        "Test Function", "Description", "Status", "Duration (s)"
    ]
    col_widths = [10, 16, 16, 30, 40, 60, 10, 12]

    ws.row_dimensions[1].height = 22
    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci)
        apply_header_style(cell)
        cell.value = h
        set_col_width(ws, ci, w)

    for ri, t in enumerate(tests, start=2):
        bg = Colors.ALT_ROW if ri % 2 == 0 else "FFFFFF"
        stat_bg, stat_fg = status_colors(t["status"])

        row_data = [
            t["tc_id"],
            t["module"],
            t["category"],
            t["class"],
            t["test_name"],
            t["description"],
            t["status"],
            t["duration_s"],
        ]
        for ci, val in enumerate(row_data, start=1):
            cell = ws.cell(row=ri, column=ci)
            cell.value = val
            if ci == 7:  # Status column
                apply_data_style(cell, bg=stat_bg, fg=stat_fg, bold=True, align="center")
            elif ci == 1:
                apply_data_style(cell, bg=bg, bold=True, align="center")
            elif ci == 8:
                apply_data_style(cell, bg=bg, align="center")
            else:
                apply_data_style(cell, bg=bg, align="left")
        ws.row_dimensions[ri].height = 16

    # Auto-filter
    ws.auto_filter.ref = f"A1:H{len(tests) + 1}"


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3: Failures Detail
# ─────────────────────────────────────────────────────────────────────────────
def write_failures_sheet(wb: openpyxl.Workbook, tests: list[dict]):
    ws = wb.create_sheet("Failure Details")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    failures = [t for t in tests if t["status"] in ("FAILED", "ERROR")]

    if not failures:
        ws.merge_cells("A1:D1")
        cell = ws["A1"]
        cell.value = "✅  No failures recorded in this test run."
        cell.fill  = make_fill(Colors.PASS_BG)
        cell.font  = Font(bold=True, color=Colors.PASS_FG, size=12, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        return

    headers = ["TC ID", "Test Function", "Status", "Failure / Error Message"]
    col_widths = [10, 45, 10, 120]
    ws.row_dimensions[1].height = 22

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci)
        apply_header_style(cell, bg=Colors.FAIL_BG.replace("FEE2E2","991B1B"), fg="FFFFFF")
        cell.value = h
        set_col_width(ws, ci, w)

    for ri, t in enumerate(failures, start=2):
        stat_bg, stat_fg = status_colors(t["status"])
        row_data = [t["tc_id"], t["test_name"], t["status"], t["failure_msg"] or "No traceback captured."]
        for ci, val in enumerate(row_data, start=1):
            cell = ws.cell(row=ri, column=ci)
            cell.value = val
            if ci == 3:
                apply_data_style(cell, bg=stat_bg, fg=stat_fg, bold=True, align="center")
            else:
                apply_data_style(cell, bg=Colors.FAIL_BG if ri % 2 == 0 else "FFF5F5",
                                  align="left")
        ws.row_dimensions[ri].height = 60


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 4: Category Pivot
# ─────────────────────────────────────────────────────────────────────────────
def write_pivot_sheet(wb: openpyxl.Workbook, tests: list[dict]):
    ws = wb.create_sheet("Category Pivot")
    ws.sheet_view.showGridLines = False

    # Group by category → module
    from collections import defaultdict
    pivot = defaultdict(lambda: defaultdict(lambda: {"total": 0, "passed": 0, "failed": 0, "skipped": 0}))
    for t in tests:
        p = pivot[t["category"]][t["module"]]
        p["total"]   += 1
        if t["status"] == "PASSED":   p["passed"]  += 1
        elif t["status"] == "FAILED": p["failed"]  += 1
        else:                          p["skipped"] += 1

    row = 1
    for cat, modules in pivot.items():
        # Category header
        ws.merge_cells(f"A{row}:F{row}")
        cat_cell = ws.cell(row=row, column=1)
        cat_cell.value = f"  🗂️  {cat}"
        cat_cell.fill  = make_fill(Colors.HEADER_BG)
        cat_cell.font  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        cat_cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 24
        row += 1

        # Column headers
        sub_headers = ["Module", "Total", "Passed", "Failed", "Skipped", "Pass Rate"]
        sub_widths   = [28, 10, 10, 10, 10, 12]
        for ci, (h, w) in enumerate(zip(sub_headers, sub_widths), start=1):
            cell = ws.cell(row=row, column=ci)
            apply_header_style(cell, bg="334155")
            cell.value = h
            set_col_width(ws, ci, w)
        ws.row_dimensions[row].height = 18
        row += 1

        for mi, (mod_name, counts) in enumerate(modules.items()):
            pct = round(counts["passed"] / counts["total"] * 100, 1) if counts["total"] else 0
            bg = Colors.ALT_ROW if mi % 2 == 0 else "FFFFFF"
            for ci, val in enumerate(
                [mod_name, counts["total"], counts["passed"], counts["failed"], counts["skipped"], f"{pct}%"],
                start=1
            ):
                cell = ws.cell(row=row, column=ci)
                cell.value = val
                apply_data_style(cell, bg=bg, align="center" if ci > 1 else "left")
            ws.row_dimensions[row].height = 16
            row += 1

        row += 1  # spacer


# ─────────────────────────────────────────────────────────────────────────────
# Main entry point
# ─────────────────────────────────────────────────────────────────────────────
def generate_report(input_json: str | None, output_xlsx: str, run_meta: dict | None = None):
    # Load test data
    if input_json and os.path.exists(input_json):
        print(f"📂  Loading pytest JSON report from: {input_json}")
        report = load_json_report(input_json)
        tests  = extract_tests(report)
        if not tests:
            print("⚠️  No tests found in JSON report, using static registry.")
            tests = build_fallback_tests()
        else:
            # Merge static descriptions for PENDING entries
            static = {t["tc_id"]: t for t in build_fallback_tests()}
            for t in tests:
                if t["tc_id"] in static and not t["description"]:
                    t["description"] = static[t["tc_id"]]["description"]
    else:
        print("ℹ️  No JSON report found — using static test-case registry (all PENDING).")
        tests = build_fallback_tests()

    if run_meta is None:
        run_meta = {}

    # Add timestamp
    run_meta.setdefault("timestamp", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"))

    # Create workbook
    wb = openpyxl.Workbook()

    print("📊  Building Summary sheet …")
    write_summary_sheet(wb, tests, run_meta)

    print("📋  Building Test Cases sheet …")
    write_testcases_sheet(wb, tests)

    print("❌  Building Failure Details sheet …")
    write_failures_sheet(wb, tests)

    print("🗂️  Building Category Pivot sheet …")
    write_pivot_sheet(wb, tests)

    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_xlsx) if os.path.dirname(output_xlsx) else ".", exist_ok=True)
    wb.save(output_xlsx)
    print(f"\n✅  Report saved → {output_xlsx}")
    print(f"    Total tests : {len(tests)}")
    passed_count = sum(1 for t in tests if t["status"] == "PASSED")
    failed_count = sum(1 for t in tests if t["status"] == "FAILED")
    print(f"    Passed      : {passed_count}")
    print(f"    Failed      : {failed_count}")
    print(f"    Pass Rate   : {round(passed_count/len(tests)*100,1) if tests else 0}%")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Digipay XLSX Test Report Generator")
    parser.add_argument("--input",  default=".test-results/report.json",
                        help="Path to pytest-json-report output JSON file")
    parser.add_argument("--output", default="reports/E2E_Test_Report_Digipay.xlsx",
                        help="Output .xlsx file path")
    parser.add_argument("--run-id",  default="")
    parser.add_argument("--sha",     default="")
    parser.add_argument("--branch",  default="main")
    parser.add_argument("--app-url", default="https://harishbalaji826-ops.github.io/Digipay-Web/")
    args = parser.parse_args()

    meta = {
        "run_id":      args.run_id  or os.environ.get("GITHUB_RUN_ID",     "local"),
        "sha":         args.sha     or os.environ.get("GITHUB_SHA",        "local"),
        "branch":      args.branch  or os.environ.get("GITHUB_REF_NAME",   "main"),
        "repo":        os.environ.get("GITHUB_REPOSITORY", "harishbalaji826-ops/Digipay-Web"),
        "app_url":     args.app_url or os.environ.get("DIGIPAY_BASE_URL",
                        "https://harishbalaji826-ops.github.io/Digipay-Web/"),
        "backend_url": os.environ.get("DIGIPAY_BACKEND_URL", "Railway"),
        "env":         "GitHub Pages + Railway",
        "python":      sys.version.split()[0],
    }

    generate_report(args.input, args.output, meta)
